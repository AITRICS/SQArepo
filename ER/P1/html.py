#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pw_to_excel.py  —  Playwright index.html → Excel Report
Step 1개 = 1행 (TC_002_001_01 ~ TC_002_005_xx)
Usage: python pw_to_excel.py <index.html> [output.xlsx]
"""
import sys, os, re, base64, zipfile, io, json, tempfile
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

KST = timezone(timedelta(hours=9))

SCENARIO_MAP = {
    "TC_002_001": "Scenario 1",
    "TC_002_002": "Scenario 2",
    "TC_002_003": "Scenario 3",
    "TC_002_004": "Scenario 4",
    "TC_002_005": "Scenario 5",
}
SCENARIO_LABEL = {
    "Scenario 1": "Scenario 1 (Landing Page)",
    "Scenario 2": "Scenario 2 (Home)",
    "Scenario 3": "Scenario 3 (Dashboard)",
    "Scenario 4": "Scenario 4 (Patient_Detail)",
    "Scenario 5": "Scenario 5 (Dashboard Options)",
}

TC_RESULT_FOLDER = {
    "TC_002_001": "tests-TC_002_001-TC-002-001-TC-002-001-Landing-Page",
    "TC_002_002": "tests-TC_002_002-TC-002-002-TC-002-002-Home",
    "TC_002_003": "tests-TC_002_003-TC-002-003-TC-002-003-Dashboard",
    "TC_002_004": "tests-TC_002_004-TC-002-004-TC-002-004-환자상세",
    "TC_002_005": "tests-TC_002_005-TC-002-005-TC-002-005-Dashboard-Options",
}

SCENARIO_ORDER = ["Scenario 1", "Scenario 2", "Scenario 3", "Scenario 4", "Scenario 5"]

def _side(c="D0D0D0"): return Side(style="thin", color=c)
def _border(c="D0D0D0"):
    s = _side(c); return Border(left=s, right=s, top=s, bottom=s)
def _fill(h): return PatternFill("solid", fgColor=h)

HEADER_FILL = _fill("2E4057")
PASS_FILL   = _fill("D6EAD6")
FAIL_FILL   = _fill("FAD7D7")
ALT_FILL    = _fill("F5F8FF")
WHITE_FILL  = _fill("FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=14)
PASS_FONT   = Font(name="Arial", bold=True, color="276227", size=10)
FAIL_FONT   = Font(name="Arial", bold=True, color="A61B28", size=10)

def _cell(ws, row, col, value="", font=None, fill=None,
          align_h="left", align_v="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BODY_FONT
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    c.border = _border()
    return c

def _merge_title(ws, row, c1, c2, val):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font = TITLE_FONT; c.fill = _fill("1A2E45")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border("888888"); ws.row_dimensions[row].height = 38

def _merge_section(ws, row, c1, c2, val):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill = _fill("3A5F8A")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border("888888"); ws.row_dimensions[row].height = 22

def ms_to_hms(ms):
    if not ms: return "0s"
    s = int(ms/1000); h,r = divmod(s,3600); m,s = divmod(r,60)
    if h: return f"{h}h {m:02d}m {s:02d}s"
    if m: return f"{m}m {s:02d}s"
    return f"{s}s"

def utc_to_kst(iso):
    if not iso: return None, ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z","+00:00")).astimezone(KST)
        return dt, dt.strftime("%Y-%m-%d %H:%M:%S")
    except: return None, ""

def tc_key_from_filename(f):
    m = re.search(r'(TC_\d{3}_\d{3})', f, re.IGNORECASE)
    return m.group(1).upper() if m else ""

def outcome_label(o):
    return {"expected":"PASS","unexpected":"FAIL","skipped":"SKIP","flaky":"FLAKY"}.get(o, o.upper())

def step_num_from_title(title):
    m = re.match(r'^(\d+)\)\s*(.*)', title.strip())
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None, title.strip()

def find_attachment_dir(test_results_root, tc_key):
    if not test_results_root or not os.path.isdir(test_results_root):
        return None
    folder = TC_RESULT_FOLDER.get(tc_key)
    if folder:
        att_dir = os.path.join(test_results_root, folder, "attachments")
        if os.path.isdir(att_dir):
            return att_dir
    tc_norm = tc_key.replace("_", "-").lower()
    for name in sorted(os.listdir(test_results_root)):
        if tc_norm in name.lower():
            att_dir = os.path.join(test_results_root, name, "attachments")
            if os.path.isdir(att_dir):
                return att_dir
    return None

def find_screenshot_in_attachments(att_dir, step_num):
    """step_num에 해당하는 스크린샷 파일 경로 리스트 반환 (여러 장 지원)"""
    if not att_dir or not os.path.isdir(att_dir):
        return []
    candidates = []
    for fname in sorted(os.listdir(att_dir)):
        is_png = (fname.lower().endswith(".png") or "-png-" in fname.lower())
        if not is_png:
            continue
        m = re.match(r'TC[-_]\d{3}[-_]\d{3}[-_](\d+)[-_]', fname, re.IGNORECASE)
        if m and int(m.group(1)) == step_num:
            candidates.append(os.path.join(att_dir, fname))
    return candidates

def make_thumb(path, max_w=260, max_h=110):
    try:
        with PILImage.open(path) as img:
            img.thumbnail((max_w, max_h), PILImage.LANCZOS)
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(tmp.name, "PNG"); return tmp.name
    except Exception as e:
        print(f"  [WARN] thumb: {e}"); return None


def make_combined_thumb(paths, max_w=260, max_h=110, gap=4):
    """여러 이미지를 세로로 합쳐 하나의 썸네일 파일로 반환"""
    try:
        thumbs = []
        for p in paths:
            with PILImage.open(p) as img:
                img2 = img.copy().convert("RGBA")
                img2.thumbnail((max_w, max_h), PILImage.LANCZOS)
                thumbs.append(img2)
        if not thumbs:
            return None
        total_h = sum(t.height for t in thumbs) + gap * (len(thumbs) - 1)
        combined = PILImage.new("RGBA", (max_w, total_h), (255, 255, 255, 255))
        y = 0
        for t in thumbs:
            combined.paste(t, (0, y))
            y += t.height + gap
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        combined.save(tmp.name, "PNG")
        return tmp.name
    except Exception as e:
        print(f"  [WARN] combined_thumb: {e}"); return None

def parse_html(html_path):
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()
    m = re.search(r'data:application/zip;base64,([A-Za-z0-9+/=\s]+)', content)
    if not m: raise RuntimeError("zip not found in index.html")
    raw = base64.b64decode(m.group(1).strip().replace("\n","").replace("\r",""))
    z   = zipfile.ZipFile(io.BytesIO(raw))
    report = json.loads(z.read("report.json").decode("utf-8"))
    fjs = {}
    for name in z.namelist():
        if name != "report.json" and name.endswith(".json"):
            d = json.loads(z.read(name).decode("utf-8"))
            fjs[d["fileId"]] = d
    return report, fjs

def collect_all(report, fjs, pw_dir):
    test_results_root = os.path.join(os.path.dirname(pw_dir), "test-results")
    if not os.path.isdir(test_results_root):
        test_results_root = os.path.join(os.path.dirname(os.path.dirname(pw_dir)), "test-results")
    print(f"      test-results root: {test_results_root}")

    result = {sg: [] for sg in SCENARIO_ORDER}

    for f_meta in report["files"]:
        fid      = f_meta["fileId"]
        tc_key   = tc_key_from_filename(f_meta["fileName"])
        scenario = SCENARIO_MAP.get(tc_key, "Unknown")
        fdata    = fjs.get(fid, {})

        att_dir  = find_attachment_dir(test_results_root, tc_key)
        print(f"      {tc_key} attachments: {att_dir}")

        for t in fdata.get("tests", []):
            res     = t["results"][0] if t.get("results") else {}
            sdt, ss = utc_to_kst(res.get("startTime",""))
            dur     = t.get("duration", 0)
            edt     = (sdt + timedelta(milliseconds=dur) if sdt else None)
            es      = edt.strftime("%Y-%m-%d %H:%M:%S") if edt else ""

            # 재귀적으로 모든 하위 step에서 숫자) 형식의 TC step만 수집
            def collect_steps(steps):
                result_steps = []
                for s in steps:
                    title = s.get("title","").strip()
                    if re.match(r'^\d+\)', title):
                        result_steps.append(s)
                    # 하위 step도 재귀 탐색
                    sub = s.get("steps", [])
                    if sub:
                        result_steps.extend(collect_steps(sub))
                return result_steps
            raw_steps = collect_steps(res.get("steps",[]))

            step_rows = []
            for i, s in enumerate(raw_steps):
                sn, test_name = step_num_from_title(s["title"])
                sn = sn or (i+1)
                img_paths = find_screenshot_in_attachments(att_dir, sn)  # 리스트
                step_rows.append({
                    "step_num":  sn,
                    "test_name": test_name,
                    "duration":  s.get("duration", 0),
                    "img_paths": img_paths,  # 리스트로 저장
                })

            if scenario in result:
                result[scenario].append({
                    "tc_id": tc_key, "title": t["title"],
                    "outcome": outcome_label(t.get("outcome","unknown")),
                    "start_dt": sdt, "start_str": ss,
                    "end_dt": edt,   "end_str": es,
                    "duration_ms": dur, "step_rows": step_rows,
                })
    return result

def build_summary(wb, all_data):
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [28,22,14,14,12,20,20,14]):
        ws.column_dimensions[col].width = w
    row = 1
    _merge_title(ws, row, 1, 8, "[ER] P1 Auto Test Report  —  Summary"); row += 1

    all_tcs = [t for tcs in all_data.values() for t in tcs]
    total = len(all_tcs)
    pn    = sum(1 for t in all_tcs if t["outcome"]=="PASS")
    fn    = sum(1 for t in all_tcs if t["outcome"]=="FAIL")
    sdts  = [t["start_dt"] for t in all_tcs if t["start_dt"]]
    edts  = [t["end_dt"]   for t in all_tcs if t["end_dt"]]
    gs    = min(sdts).strftime("%Y-%m-%d %H:%M:%S") if sdts else ""
    ge    = max(edts).strftime("%Y-%m-%d %H:%M:%S") if edts else ""
    gd    = int((max(edts)-min(sdts)).total_seconds()) if (sdts and edts) else ""

    _merge_section(ws, row, 1, 8, "◆  Total Summary"); row += 1
    for lb, val in [("Total TestCases",total),("PASS",pn),("FAIL",fn),
                    ("Start Time",gs),("End Time",ge),("Duration (sec)",gd)]:
        ws.row_dimensions[row].height = 20
        _cell(ws, row, 1, lb, font=Font(name="Arial", bold=True, size=10), fill=_fill("EEF3FB"))
        c = _cell(ws, row, 2, val, fill=WHITE_FILL)
        if lb=="PASS": c.font=PASS_FONT
        elif lb=="FAIL": c.font=FAIL_FONT
        ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8); row += 1

    row += 1
    _merge_section(ws, row, 1, 8, "◆  Scenario Summary"); row += 1
    for h,ci in zip(["Scenario","TestCases","PASS","FAIL","Pass Rate",
                     "Start Time","End Time","Duration(sec)"], range(1,9)):
        _cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align_h="center")
    ws.row_dimensions[row].height = 20; row += 1

    for sg in SCENARIO_ORDER:
        tcs  = all_data.get(sg, [])
        t_   = len(tcs); p_ = sum(1 for t in tcs if t["outcome"]=="PASS")
        f_   = sum(1 for t in tcs if t["outcome"]=="FAIL")
        sl   = [t["start_dt"] for t in tcs if t["start_dt"]]
        el   = [t["end_dt"]   for t in tcs if t["end_dt"]]
        ss   = min(sl).strftime("%Y-%m-%d %H:%M:%S") if sl else "-"
        se   = max(el).strftime("%Y-%m-%d %H:%M:%S") if el else "-"
        sd   = int((max(el)-min(sl)).total_seconds()) if (sl and el) else 0
        rate = f"{p_/t_*100:.0f}%" if t_ else "-"
        label = SCENARIO_LABEL.get(sg, sg)
        fill  = ALT_FILL if row%2==0 else WHITE_FILL
        for ci, val in enumerate([label,t_,p_,f_,rate,ss,se,sd], 1):
            c = _cell(ws, row, ci, val, fill=fill, align_h="center")
            if ci==1: c.alignment=Alignment(horizontal="left",vertical="center")
            if ci==3: c.font=PASS_FONT
            if ci==4: c.font=FAIL_FONT
        ws.row_dimensions[row].height = 20; row += 1

def build_scenario_sheet(wb, sg_key, tc_list, pw_dir):
    label = SCENARIO_LABEL.get(sg_key, sg_key)
    ws    = wb.create_sheet(title=label[:31])
    ws.sheet_view.showGridLines = False

    COLS = [
        ("TC ID",       18),
        ("Test Name",   36),
        ("Result",      10),
        ("Start Time",  20),
        ("End Time",    20),
        ("Duration",    12),
        ("Screenshot",  48),
    ]
    for ci,(_, w) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 1
    _merge_title(ws, row, 1, len(COLS), f"[ER] P1 Auto Test Report  —  {label}"); row += 1
    for ci,(h,_) in enumerate(COLS,1):
        _cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align_h="center")
    ws.row_dimensions[row].height = 22; row += 1

    thumbs = []
    IMG_H  = 130

    for tc in tc_list:
        tc_key    = tc["tc_id"]
        outcome   = tc["outcome"]
        step_rows = tc["step_rows"]
        cur_dt    = tc["start_dt"]

        for sr in step_rows:
            sn        = sr["step_num"]
            tc_id_s   = f"{tc_key}_{sn:02d}"
            s_str     = cur_dt.strftime("%Y-%m-%d %H:%M:%S") if cur_dt else ""
            dur_ms    = sr["duration"]
            if cur_dt and dur_ms:
                ndt   = cur_dt + timedelta(milliseconds=dur_ms)
                e_str = ndt.strftime("%Y-%m-%d %H:%M:%S")
            else:
                ndt = None; e_str = ""

            fill  = ALT_FILL if row%2==0 else WHITE_FILL
            rfill = PASS_FILL if outcome=="PASS" else (FAIL_FILL if outcome=="FAIL" else WHITE_FILL)

            img_paths = sr["img_paths"]  # 리스트

            # 이미지 수에 맞게 행 높이 설정
            img_count = max(len(img_paths), 1)
            ws.row_dimensions[row].height = IMG_H * img_count

            _cell(ws, row, 1, tc_id_s, fill=fill, align_v="top",
                  font=Font(name="Arial", bold=True, size=10))
            _cell(ws, row, 2, sr["test_name"], fill=fill, wrap=True, align_v="top")
            c = _cell(ws, row, 3, outcome, fill=rfill, align_h="center", align_v="top")
            c.font = PASS_FONT if outcome=="PASS" else (FAIL_FONT if outcome=="FAIL" else BODY_FONT)
            _cell(ws, row, 4, s_str,             fill=fill, align_h="center", align_v="top")
            _cell(ws, row, 5, e_str,             fill=fill, align_h="center", align_v="top")
            _cell(ws, row, 6, ms_to_hms(dur_ms), fill=fill, align_h="center", align_v="top")
            _cell(ws, row, 7, "",                fill=fill)

            # 여러 이미지를 세로로 합쳐서 하나의 이미지로 삽입
            if img_paths:
                combined = make_combined_thumb(img_paths)
                if combined:
                    thumbs.append(combined)
                    try:
                        xl_img = XLImage(combined)
                        ws.add_image(xl_img, f"{get_column_letter(7)}{row}")
                    except Exception as e:
                        print(f"  [WARN] img insert: {e}")

            row += 1
            cur_dt = ndt

        # TC 구분 빈 행
        for ci in range(1, len(COLS)+1):
            c = ws.cell(row=row, column=ci)
            c.fill = _fill("E8ECF5"); c.border = _border("C0C8DC")
        ws.row_dimensions[row].height = 6; row += 1

    return thumbs

def generate(html_path, out_xlsx=None):
    if not os.path.exists(html_path):
        print(f"[ERROR] Not found: {html_path}"); return
    pw_dir = os.path.dirname(os.path.abspath(html_path))
    if not out_xlsx:
        out_xlsx = os.path.join(pw_dir, "pw_report.xlsx")

    print(f"[1/3] Parsing {html_path} ...")
    report, fjs = parse_html(html_path)

    print(f"[2/3] Collecting data ...")
    all_data = collect_all(report, fjs, pw_dir)
    for sg, tcs in all_data.items():
        steps_total = sum(len(t["step_rows"]) for t in tcs)
        print(f"      {sg}: {len(tcs)} TC, {steps_total} step rows")

    print(f"[3/3] Building Excel ...")
    wb = Workbook()
    build_summary(wb, all_data)
    all_thumbs = []
    for sg in SCENARIO_ORDER:
        t = build_scenario_sheet(wb, sg, all_data.get(sg,[]), pw_dir)
        all_thumbs.extend(t)
    wb.save(out_xlsx)
    print(f"[OK] Saved: {out_xlsx}")
    for p in all_thumbs:
        try: os.unlink(p)
        except: pass

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python pw_to_excel.py <index.html> [output.xlsx]"); sys.exit(1)
    generate(sys.argv[1], sys.argv[2] if len(sys.argv)>=3 else None)
