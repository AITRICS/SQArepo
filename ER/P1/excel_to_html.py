#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_to_html.py  —  Excel Report → HTML Report

Usage:
  python excel_to_html.py <pw_report.xlsx> [output.html]

  스크린샷은 xlsx 기준 상위 폴더의 test-results/.../attachments/ 에서 자동 탐색.
  환경변수 TEST_RESULTS_DIR 로 직접 지정 가능.
"""
import sys, os, re, json, base64
from datetime import datetime
from openpyxl import load_workbook

REPORT_TITLE = "[ER] P1 Auto Test Report"

SCENARIO_SHEETS = [
    "Scenario 1 (Landing Page)",
    "Scenario 2 (Home)",
    "Scenario 3 (Dashboard)",
    "Scenario 4 (Patient_Detail)",
    "Scenario 5 (Dashboard Options)",
]
SCENARIO_SHORT = {
    "Scenario 1 (Landing Page)":      "Scenario 1",
    "Scenario 2 (Home)":              "Scenario 2",
    "Scenario 3 (Dashboard)":         "Scenario 3",
    "Scenario 4 (Patient_Detail)":    "Scenario 4",
    "Scenario 5 (Dashboard Options)": "Scenario 5",
}
TC_RESULT_FOLDER = {
    "TC_002_001": "tests-TC_002_001-TC-002-001-TC-002-001-Landing-Page",
    "TC_002_002": "tests-TC_002_002-TC-002-002-TC-002-002-Home",
    "TC_002_003": "tests-TC_002_003-TC-002-003-TC-002-003-Dashboard",
    "TC_002_004": "tests-TC_002_004-TC-002-004-TC-002-004-환자상세",
    "TC_002_005": "tests-TC_002_005-TC-002-005-TC-002-005-Dashboard-Options",
}
SHEET_TO_TC = {
    "Scenario 1 (Landing Page)":      "TC_002_001",
    "Scenario 2 (Home)":              "TC_002_002",
    "Scenario 3 (Dashboard)":         "TC_002_003",
    "Scenario 4 (Patient_Detail)":    "TC_002_004",
    "Scenario 5 (Dashboard Options)": "TC_002_005",
}

def he(s):
    if s is None: s = ""
    return (str(s).replace("&","&amp;").replace("<","&lt;")
            .replace(">","&gt;").replace('"',"&quot;"))

def safe(v):
    return "" if v is None else str(v).strip()

def img_b64(path):
    fname_lower = os.path.basename(path).lower()
    if "-png-" in fname_lower or fname_lower.endswith(".png"):
        mime = "image/png"
    else:
        mime = "image/jpeg"
    with open(path, "rb") as f:
        return f"data:{mime};base64,{base64.b64encode(f.read()).decode()}"

def find_attachment_dir(test_results_root, tc_key):
    if not test_results_root or not os.path.isdir(test_results_root):
        return None
    folder = TC_RESULT_FOLDER.get(tc_key)
    if folder:
        att = os.path.join(test_results_root, folder, "attachments")
        if os.path.isdir(att):
            return att
    # fallback: 폴더명에 tc_key 포함
    tc_norm = tc_key.replace("_", "-").lower()
    for name in sorted(os.listdir(test_results_root)):
        if tc_norm in name.lower():
            att = os.path.join(test_results_root, name, "attachments")
            if os.path.isdir(att):
                return att
    return None

def find_screenshot(att_dir, step_num):
    """step_num에 해당하는 스크린샷 파일 경로 리스트 반환 (여러 장 지원)"""
    if not att_dir or not os.path.isdir(att_dir):
        return []
    candidates = []
    for fname in sorted(os.listdir(att_dir)):
        is_png = (fname.lower().endswith(".png") or "-png-" in fname.lower())
        if not is_png:
            continue
        m = re.match(r'TC[-_]\d{3}[-_]\d{3}[-_](\d+)[-_]', fname, re.IGNORECASE)
        if m and int(m.group(1)) == step_num:
            candidates.append(os.path.join(att_dir, fname))
    return candidates

def step_num_from_tc_id(tc_id_str):
    """TC_002_003_14 → 14"""
    m = re.search(r'_(\d+)$', safe(tc_id_str))
    return int(m.group(1)) if m else None

def tc_base_from_tc_id(tc_id_str):
    """TC_002_003_14 → TC_002_003"""
    m = re.match(r'(TC_\d{3}_\d{3})', safe(tc_id_str), re.IGNORECASE)
    return m.group(1).upper() if m else None

def read_scenario_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, r in enumerate(rows):
        if r and any(safe(c) == "TC ID" for c in r):
            header_row = i; break
    if header_row is None:
        return [], []
    headers   = [safe(c) for c in rows[header_row]]
    data_rows = []
    for r in rows[header_row+1:]:
        if not r or all(c is None for c in r): continue
        data_rows.append({h: safe(v) for h, v in zip(headers, r)})
    return headers, data_rows

CSS = """
*{box-sizing:border-box;}
body{margin:0;background:#f4f6fb;color:#1f2330;
     font-family:system-ui,-apple-system,'Segoe UI',Roboto,'Noto Sans KR',Arial,sans-serif;
     font-size:15px;}
.wrap{max-width:none;margin:0 auto;padding:24px 36px;}
.page-title{font-size:26px;font-weight:800;letter-spacing:-.3px;margin:0 0 18px 0;color:#1a2e45;}
.tabs{display:flex;gap:8px;flex-wrap:nowrap;overflow-x:auto;margin:0 0 18px 0;padding-bottom:4px;}
.tabBtn{border:1.5px solid #c9d4e8;background:#f9faff;color:#374151;
        border-radius:7px;padding:10px 18px;font-weight:700;font-size:13px;
        cursor:pointer;white-space:nowrap;transition:all .15s;}
.tabBtn:hover{background:#eef1ff;border-color:#8a9fd4;}
.tabBtn.active{background:#2e4057;border-color:#2e4057;color:#fff;}
.tabPanel{display:none;}
.tabPanel.active{display:block;}
.card{background:#fff;border-radius:12px;padding:28px 28px 22px;
      margin-bottom:20px;border:1px solid #e4e8f2;
      box-shadow:0 2px 14px rgba(31,35,48,.05);}
.card-title{font-size:20px;font-weight:800;color:#1a2e45;margin:0 0 18px 0;}
.section-label{font-size:13px;font-weight:700;color:#6b7280;
               text-transform:uppercase;letter-spacing:.6px;margin:22px 0 10px 0;}
.summary-table{width:100%;border-collapse:collapse;font-size:14px;}
.summary-table th,.summary-table td{padding:11px 16px;border:1px solid #e4e8f2;
                                     text-align:left;vertical-align:middle;}
.summary-table th{background:#f0f4fb;font-weight:700;color:#374151;width:220px;}
.summary-table td{background:#fff;}
.sg-table{width:100%;border-collapse:collapse;font-size:13px;table-layout:fixed;}
.sg-table th{background:#2e4057;color:#fff;font-weight:700;
             padding:11px 14px;border:1px solid #3a5070;
             position:sticky;top:0;z-index:10;text-align:left;}
.sg-table td{padding:8px 12px;border:1px solid #e4e8f2;vertical-align:middle;word-break:break-word;}
.sg-table tr:nth-child(even) td{background:#f7f9ff;}
.sg-table tr:nth-child(odd)  td{background:#fff;}
.sg-table tr:hover td{background:#eef2fb;}
.badge{display:inline-block;padding:3px 12px;border-radius:999px;font-weight:700;font-size:12px;}
.pass{background:#d6ead6;color:#276227;}
.fail{background:#fad7d7;color:#a61b28;}
.skip{background:#f0f0f0;color:#555;}
.ss-wrap{display:flex;flex-wrap:wrap;gap:6px;align-items:flex-start;}
.ss-img{max-width:300px;max-height:200px;border-radius:6px;border:1px solid #dde3f0;
        cursor:pointer;transition:transform .15s;object-fit:contain;background:#f9f9f9;display:block;}
.ss-img:hover{transform:scale(1.03);}
.no-img{color:#bbb;font-size:12px;}
.stat-row{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:20px;}
.stat-pill{background:#f0f4fb;border-radius:10px;padding:14px 22px;
           border:1px solid #dde3f0;min-width:140px;}
.stat-pill .sp-label{font-size:12px;color:#6b7280;font-weight:600;}
.stat-pill .sp-value{font-size:28px;font-weight:800;color:#1a2e45;line-height:1.2;}
.stat-pill.pass-pill .sp-value{color:#276227;}
.stat-pill.fail-pill .sp-value{color:#a61b28;}
.charts-row{display:flex;gap:24px;flex-wrap:wrap;margin-top:16px;}
.chart-box{flex:1;min-width:300px;max-width:520px;background:#f9faff;
           border-radius:10px;padding:18px;border:1px solid #e4e8f2;}
.chart-box h3{margin:0 0 12px 0;font-size:14px;color:#374151;}
.chart-box canvas{width:100% !important;height:300px !important;}
"""

def build_html(wb, out_html, test_results_root):
    lines = []; A = lines.append

    A("<!DOCTYPE html><html><head><meta charset='UTF-8'>")
    A("<script src='https://cdn.jsdelivr.net/npm/chart.js'></script>")
    A(f"<style>{CSS}</style>")
    A(f"<title>{he(REPORT_TITLE)}</title></head><body><div class='wrap'>")
    A(f"<div class='page-title'>{he(REPORT_TITLE)}</div>")

    # 탭 버튼
    tab_keys   = ["summary"] + [f"sg_{i}" for i in range(len(SCENARIO_SHEETS))]
    tab_labels = ["📋 Test Results"] + [f"🔬 {SCENARIO_SHORT.get(s,s)}" for s in SCENARIO_SHEETS]
    A("<div class='tabs' id='topTabs'>")
    for i,(k,lb) in enumerate(zip(tab_keys, tab_labels)):
        active = " active" if i==0 else ""
        A(f"<button class='tabBtn{active}' data-tab='{k}'>{he(lb)}</button>")
    A("</div>")

    # ── Summary 패널 ──────────────────────────────────────────
    A("<div class='tabPanel active' id='summary'><div class='card'>")
    A("<div class='card-title'>Test Summary</div>")

    scenarios_data = {}
    for sname in SCENARIO_SHEETS:
        hdr, rows = read_scenario_sheet(wb, sname)
        scenarios_data[sname] = (hdr, rows)

    all_rows   = [r for _,rows in scenarios_data.values() for r in rows]
    total      = len([r for r in all_rows if r.get("TC ID","")])
    total_scen = len([s for s in SCENARIO_SHEETS if scenarios_data[s][1]])
    pass_n     = sum(1 for r in all_rows if r.get("Result","").upper()=="PASS")
    fail_n     = sum(1 for r in all_rows if r.get("Result","").upper()=="FAIL")

    starts  = [r.get("Start Time","") for r in all_rows if r.get("Start Time","")]
    ends    = [r.get("End Time","")   for r in all_rows if r.get("End Time","")]
    g_start = min(starts) if starts else ""
    g_end   = max(ends)   if ends   else ""
    try:
        d_sec = int((datetime.strptime(g_end,"%Y-%m-%d %H:%M:%S") -
                     datetime.strptime(g_start,"%Y-%m-%d %H:%M:%S")).total_seconds())
    except: d_sec = ""

    A("<div class='stat-row'>")
    for lb,val,cls in [("Total Scenarios", total_scen, ""),
                       ("Total Test Cases", total, ""),
                       ("PASS", pass_n, "pass-pill"),
                       ("FAIL", fail_n, "fail-pill")]:
        A(f"<div class='stat-pill {cls}'><div class='sp-label'>{lb}</div><div class='sp-value'>{val}</div></div>")
    A("</div>")

    A("<div class='section-label'>Overall</div>")
    A("<table class='summary-table'><tbody>")
    for k,v in [("Start Time",g_start),("End Time",g_end),("Duration (sec)",d_sec)]:
        A(f"<tr><th>{he(k)}</th><td>{he(str(v))}</td></tr>")
    A("</tbody></table>")

    A("<div class='section-label'>Scenario Breakdown</div>")
    A("<table class='summary-table'><thead><tr>")
    for h in ["Scenario","TC Count","PASS","FAIL","Pass Rate","Start","End","Duration(s)"]:
        A(f"<th>{he(h)}</th>")
    A("</tr></thead><tbody>")

    chart_labels, chart_pass, chart_fail = [], [], []
    for sname in SCENARIO_SHEETS:
        _, rows = scenarios_data[sname]
        sg_t = len([r for r in rows if r.get("TC ID","")])
        sg_p = sum(1 for r in rows if r.get("Result","").upper()=="PASS")
        sg_f = sum(1 for r in rows if r.get("Result","").upper()=="FAIL")
        sg_starts = [r.get("Start Time","") for r in rows if r.get("Start Time","")]
        sg_ends   = [r.get("End Time","")   for r in rows if r.get("End Time","")]
        s_s = min(sg_starts) if sg_starts else ""
        s_e = max(sg_ends)   if sg_ends   else ""
        try:
            s_d = int((datetime.strptime(s_e,"%Y-%m-%d %H:%M:%S") -
                       datetime.strptime(s_s,"%Y-%m-%d %H:%M:%S")).total_seconds())
        except: s_d = ""
        rate  = f"{sg_p/sg_t*100:.0f}%" if sg_t else "-"
        short = SCENARIO_SHORT.get(sname, sname)
        A(f"<tr><td>{he(sname)}</td><td style='text-align:center'>{sg_t}</td>")
        A(f"<td style='text-align:center;color:#276227;font-weight:700'>{sg_p}</td>")
        A(f"<td style='text-align:center;color:#a61b28;font-weight:700'>{sg_f}</td>")
        A(f"<td style='text-align:center'>{rate}</td>")
        A(f"<td>{he(s_s)}</td><td>{he(s_e)}</td><td style='text-align:center'>{he(str(s_d))}</td></tr>")
        chart_labels.append(short); chart_pass.append(sg_p); chart_fail.append(sg_f)
    A("</tbody></table>")

    A("<div class='section-label'>Visualizations</div>")
    A("<div class='charts-row'>")
    A("<div class='chart-box'><h3>PASS / FAIL per Scenario</h3><canvas id='barChart'></canvas></div>")
    A("<div class='chart-box'><h3>Overall PASS vs FAIL</h3><canvas id='pieChart'></canvas></div>")
    A("</div></div></div>")

    A(f"""<script>(function(){{
  function init(){{
    if(window.__chartsOk)return; window.__chartsOk=true;
    new Chart(document.getElementById('barChart'),{{type:'bar',data:{{
      labels:{json.dumps(chart_labels,ensure_ascii=False)},
      datasets:[
        {{label:'PASS',data:{json.dumps(chart_pass)},backgroundColor:'#4a90d9'}},
        {{label:'FAIL',data:{json.dumps(chart_fail)},backgroundColor:'#d94a4a'}}
      ]}},options:{{responsive:true,maintainAspectRatio:false,
        scales:{{x:{{stacked:true}},y:{{stacked:true,ticks:{{stepSize:1}}}}}},
        plugins:{{legend:{{position:'bottom'}}}}}}
    }});
    new Chart(document.getElementById('pieChart'),{{type:'doughnut',data:{{
      labels:['PASS','FAIL'],
      datasets:[{{data:[{pass_n},{fail_n}],backgroundColor:['#4a90d9','#d94a4a'],borderWidth:2}}]}},
      options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom'}}}}}}
    }});
  }}
  window.__initCharts=init;
  window.addEventListener('load',init);
}})();</script>""")

    # ── Scenario 패널들 ────────────────────────────────────────
    for si, sname in enumerate(SCENARIO_SHEETS):
        tab_id  = f"sg_{si}"
        tc_key  = SHEET_TO_TC.get(sname, "")
        att_dir = find_attachment_dir(test_results_root, tc_key)
        print(f"  [{sname}] attachments: {att_dir}")

        A(f"<div class='tabPanel' id='{tab_id}'><div class='card'>")
        A(f"<div class='card-title'>{he(sname)}</div>")

        _, rows = scenarios_data[sname]
        if not rows:
            A("<div style='color:#6b7280;padding:20px 0;'>No data.</div></div></div>"); continue

        col_widths = {
            "TC ID":      "140px",
            "Test Name":  "240px",
            "Result":     "80px",
            "Start Time": "150px",
            "End Time":   "150px",
            "Duration":   "90px",
            "Screenshot": "320px",
        }
        display_cols = ["TC ID","Test Name","Result","Start Time","End Time","Duration","Screenshot"]

        A("<div style='overflow-x:auto;'><table class='sg-table'><colgroup>")
        for col in display_cols:
            w = col_widths.get(col, "auto")
            A(f"<col style='width:{w}' />")
        A("</colgroup><thead><tr>")
        for col in display_cols:
            A(f"<th>{he(col)}</th>")
        A("</tr></thead><tbody>")

        for r in rows:
            outcome   = r.get("Result","").upper()
            badge_cls = "pass" if outcome=="PASS" else ("fail" if outcome=="FAIL" else "skip")
            tc_id_str = r.get("TC ID","")
            step_num  = step_num_from_tc_id(tc_id_str)

            A("<tr>")
            for col in display_cols:
                val = r.get(col,"")
                if col == "Result":
                    A(f"<td style='text-align:center'><span class='badge {badge_cls}'>{he(outcome or val)}</span></td>")
                elif col == "Screenshot":
                    img_paths = find_screenshot(att_dir, step_num) if (att_dir and step_num) else []
                    if img_paths:
                        A("<td><div class='ss-wrap'>")
                        for img_path in img_paths:
                            if os.path.exists(img_path):
                                try:
                                    b64   = img_b64(img_path)
                                    fname = he(os.path.basename(img_path))
                                    A(f"<img class='ss-img' src='{b64}' title='{fname}' onclick=\"window.open(this.src)\" />")
                                except Exception as e:
                                    A(f"<span class='no-img'>이미지 오류</span>")
                        A("</div></td>")
                    else:
                        A("<td class='no-img' style='text-align:center'>—</td>")
                else:
                    A(f"<td>{he(val)}</td>")
            A("</tr>")

        A("</tbody></table></div></div></div>")

    # 탭 전환 JS
    A("""<script>(function(){
  var btns=document.querySelectorAll('#topTabs .tabBtn');
  var panels=document.querySelectorAll('.tabPanel');
  btns.forEach(function(btn){
    btn.addEventListener('click',function(){
      btns.forEach(function(b){b.classList.remove('active');});
      panels.forEach(function(p){p.classList.remove('active');});
      btn.classList.add('active');
      var panel=document.getElementById(btn.getAttribute('data-tab'));
      if(panel)panel.classList.add('active');
      if(btn.getAttribute('data-tab')==='summary'&&window.__initCharts)
        setTimeout(window.__initCharts,100);
    });
  });
})();</script>""")
    A("</div></body></html>")

    with open(out_html,"w",encoding="utf-8",newline="\n") as f:
        f.write("\n".join(lines))
    print(f"[OK] HTML saved: {out_html}")

def generate(xlsx_path, out_html=None):
    if not os.path.exists(xlsx_path):
        print(f"[ERROR] Not found: {xlsx_path}"); return
    if not out_html:
        out_html = xlsx_path.replace(".xlsx","_report.html")

    xlsx_dir = os.path.dirname(os.path.abspath(xlsx_path))
    test_results_root = os.environ.get("TEST_RESULTS_DIR","")
    if not test_results_root:
        candidate = os.path.join(os.path.dirname(xlsx_dir), "test-results")
        if os.path.isdir(candidate):
            test_results_root = candidate
        else:
            candidate = os.path.normpath(os.path.join(xlsx_dir, "..", "test-results"))
            if os.path.isdir(candidate):
                test_results_root = candidate

    print(f"[INFO] test-results: {test_results_root or '(not found)'}")
    print(f"[1/2] Loading {xlsx_path} ...")
    wb = load_workbook(xlsx_path, data_only=True)
    print(f"[2/2] Building HTML ...")
    build_html(wb, out_html, test_results_root)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_html.py <pw_report.xlsx> [output.html]")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2] if len(sys.argv)>=3 else None)
